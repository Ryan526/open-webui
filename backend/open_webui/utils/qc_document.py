import base64
import io
import logging
import math
from typing import Optional

from PIL import Image, ImageDraw, ImageFont

log = logging.getLogger(__name__)


SEVERITY_COLORS = {
    "critical": (220, 38, 38),    # red
    "major": (234, 88, 12),       # orange
    "minor": (202, 138, 4),       # yellow
    "info": (37, 99, 235),        # blue
}


def find_text_on_page(
    pdf_path: str, page_number: int, search_text: str
) -> list[dict]:
    """
    Search for text on a specific PDF page and return normalized 0-1 coordinates.

    Args:
        pdf_path: Path to the PDF file.
        page_number: 1-based page number.
        search_text: Text string to search for on the page.

    Returns:
        List of dicts with {x, y, width, height} normalized to 0-1.
        Empty list if text not found.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        log.error("PyMuPDF (fitz) is not installed")
        return []

    try:
        doc = fitz.open(pdf_path)
        if page_number < 1 or page_number > len(doc):
            doc.close()
            return []

        page = doc.load_page(page_number - 1)
        page_rect = page.rect
        pw = page_rect.width
        ph = page_rect.height

        if pw == 0 or ph == 0:
            doc.close()
            return []

        matches = page.search_for(search_text)
        doc.close()

        if not matches:
            return []

        padding = 3.0  # points of padding around each match
        results = []
        for rect in matches:
            x0 = max(0.0, rect.x0 - padding) / pw
            y0 = max(0.0, rect.y0 - padding) / ph
            x1 = min(pw, rect.x1 + padding) / pw
            y1 = min(ph, rect.y1 + padding) / ph
            results.append({
                "x": x0,
                "y": y0,
                "width": x1 - x0,
                "height": y1 - y0,
            })
        return results

    except Exception as e:
        log.warning(f"Error searching text '{search_text}' on page {page_number}: {e}")
        return []


def extract_page_text(pdf_path: str, page_number: int) -> dict:
    """
    Extract text content and structure from a single PDF page using PyMuPDF.

    Args:
        pdf_path: Path to the PDF file.
        page_number: 1-based page number.

    Returns:
        Dict with:
          - raw_text: Full extracted text as a string.
          - blocks: List of text blocks with position info
            [{text, x0, y0, x1, y1}] normalized to 0-1.
          - tables: List of tables found on the page, each as a list of rows.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        log.error("PyMuPDF (fitz) is not installed")
        return {"raw_text": "", "blocks": [], "tables": []}

    try:
        doc = fitz.open(pdf_path)
        if page_number < 1 or page_number > len(doc):
            doc.close()
            return {"raw_text": "", "blocks": [], "tables": []}

        page = doc.load_page(page_number - 1)
        page_rect = page.rect
        pw = page_rect.width
        ph = page_rect.height

        # Extract raw text
        raw_text = page.get_text("text") or ""

        # Extract positioned text blocks
        blocks = []
        if pw > 0 and ph > 0:
            text_dict = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)
            for block in text_dict.get("blocks", []):
                if block.get("type") != 0:  # 0 = text block
                    continue
                block_text = ""
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        block_text += span.get("text", "")
                    block_text += "\n"
                block_text = block_text.strip()
                if not block_text:
                    continue
                bbox = block.get("bbox", (0, 0, 0, 0))
                blocks.append({
                    "text": block_text,
                    "x0": bbox[0] / pw,
                    "y0": bbox[1] / ph,
                    "x1": bbox[2] / pw,
                    "y1": bbox[3] / ph,
                })

        # Extract tables
        tables = []
        try:
            page_tables = page.find_tables()
            for table in page_tables:
                rows = table.extract()
                if rows:
                    tables.append(rows)
        except Exception as e:
            log.debug(f"Table extraction failed on page {page_number}: {e}")

        doc.close()
        return {"raw_text": raw_text, "blocks": blocks, "tables": tables}

    except Exception as e:
        log.warning(f"Error extracting text from page {page_number}: {e}")
        return {"raw_text": "", "blocks": [], "tables": []}


def convert_pdf_to_pages(file_path: str, dpi: int = 300) -> list[bytes]:
    """Convert each page of a PDF to PNG bytes using PyMuPDF."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        log.error("PyMuPDF (fitz) is not installed. Install with: pip install pymupdf")
        raise

    doc = fitz.open(file_path)
    pages = []
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap(matrix=matrix)
        png_bytes = pix.tobytes("png")
        pages.append(png_bytes)

    doc.close()
    return pages


def render_annotated_page(
    page_image_bytes: bytes,
    findings: list[dict],
) -> bytes:
    """
    Draw numbered annotations on a copy of the page image.

    Each finding dict should have:
      - finding_number: int
      - severity: str (critical|major|minor|info)
      - location: dict with {x, y, width, height} normalized 0-1
    """
    img = Image.open(io.BytesIO(page_image_bytes)).convert("RGBA")
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    img_w, img_h = img.size

    try:
        font_size = max(16, int(img_h * 0.018))
        font = ImageFont.truetype("arial.ttf", font_size)
    except (OSError, IOError):
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
        except (OSError, IOError):
            font = ImageFont.load_default()

    for finding in findings:
        loc = finding.get("location")
        if not loc:
            continue

        # Normalize location to list of rects (handles both single dict and array)
        if isinstance(loc, list):
            rects = loc
        else:
            rects = [loc]

        number = finding.get("finding_number", 0)
        severity = finding.get("severity", "info")
        color = SEVERITY_COLORS.get(severity, SEVERITY_COLORS["info"])

        for idx, rect in enumerate(rects):
            x = rect.get("x", 0) * img_w
            y = rect.get("y", 0) * img_h
            w = rect.get("width", 0.05) * img_w
            h = rect.get("height", 0.05) * img_h

            # Draw semi-transparent rectangle
            rect_color = color + (60,)  # alpha=60
            draw.rectangle([x, y, x + w, y + h], fill=rect_color, outline=color + (200,), width=2)

            # Draw numbered circle only at the first rect
            if idx == 0:
                label = str(number)
                circle_r = max(12, font_size)
                cx = x - circle_r * 0.3
                cy = y - circle_r * 0.3
                # Clamp to image bounds
                cx = max(circle_r, min(img_w - circle_r, cx))
                cy = max(circle_r, min(img_h - circle_r, cy))

                # White circle background with colored border
                draw.ellipse(
                    [cx - circle_r, cy - circle_r, cx + circle_r, cy + circle_r],
                    fill=(255, 255, 255, 230),
                    outline=color + (255,),
                    width=2,
                )

                # Number text centered in circle
                bbox = draw.textbbox((0, 0), label, font=font)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                draw.text(
                    (cx - tw / 2, cy - th / 2),
                    label,
                    fill=color + (255,),
                    font=font,
                )

    # Composite overlay onto image
    result = Image.alpha_composite(img, overlay).convert("RGB")

    buf = io.BytesIO()
    result.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def image_to_base64(image_bytes: bytes) -> str:
    """Convert image bytes to base64 data URL."""
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    return f"data:image/png;base64,{b64}"


def process_excel_for_qc(file_path: str) -> list[dict]:
    """Extract sheet data from Excel file for text-based QC analysis."""
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed, cannot process Excel files")
        return []

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        sheets.append({
            "sheet_name": sheet_name,
            "rows": rows,
            "row_count": len(rows),
        })
    wb.close()
    return sheets


def process_docx_for_qc(file_path: str) -> str:
    """Extract text content from DOCX file."""
    try:
        import docx
    except ImportError:
        log.warning("python-docx not installed, cannot process DOCX files")
        return ""

    doc = docx.Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)
